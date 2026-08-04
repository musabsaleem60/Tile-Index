"""
Seed tile_sizes and tile_rates from the client workbook's Formulas sheet.

Dry-run:
    python -m scripts.seed_tile_formulas path/to/Tiles_Inventory_Final.xlsx

Commit:
    python -m scripts.seed_tile_formulas path/to/Tiles_Inventory_Final.xlsx --commit
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path

from sqlalchemy import select

from app.db.session import SessionLocal
from app.models.entities import TileRate, TileSize


EXPECTED_SIZE_ROWS = 15
EXPECTED_RATE_ROWS = 45
GRADE_MAP = {
    "G1": "G1 Prime",
    "G2": "G2 Standard",
    "G3": "G3 Regular",
    "G1 Prime": "G1 Prime",
    "G2 Standard": "G2 Standard",
    "G3 Regular": "G3 Regular",
    "Grade 1 (Prime)": "G1 Prime",
    "Grade 2 (Standard)": "G2 Standard",
    "Grade 3 (Regular)": "G3 Regular",
}


@dataclass(frozen=True)
class ParsedSize:
    tile_size: str
    pieces_per_box: int
    area_per_box: float


@dataclass(frozen=True)
class ParsedRate:
    tile_size: str
    grade: str
    rate_per_meter: float


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", default="Formulas")
    parser.add_argument("--commit", action="store_true")
    args = parser.parse_args()

    sizes, rates = parse_workbook(args.workbook, args.sheet)
    if len(sizes) != EXPECTED_SIZE_ROWS:
        raise SystemExit(f"Expected {EXPECTED_SIZE_ROWS} size rows, found {len(sizes)}")
    if len(rates) != EXPECTED_RATE_ROWS:
        raise SystemExit(f"Expected {EXPECTED_RATE_ROWS} rate rows, found {len(rates)}")

    with SessionLocal() as db:
        diff = build_diff(db, sizes, rates)
        print(json.dumps(diff, indent=2, ensure_ascii=False))
        if not args.commit:
            print("DRY RUN ONLY. Re-run with --commit to write these changes.")
            return

        apply_changes(db, sizes, rates)
        db.commit()
        print("Committed tile size and rate card seed.")


def parse_workbook(path: Path, sheet_name: str) -> tuple[list[ParsedSize], list[ParsedRate]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required. Install backend requirements first.") from exc

    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")

    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")

    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    size_headers = {
        "tile_size": ("tile sizes", "tile size"),
        "pieces_per_box": ("piece per box", "pieces per box"),
        "area_per_box": ("area per box",),
    }
    rate_headers = {
        "grade": ("grade",),
        "tile_size": ("tiles size", "tile size"),
        "rate_per_meter": ("rate per meter",),
    }

    sizes = _parse_table(rows, size_headers, _size_from_row)
    rates = _parse_table(rows, rate_headers, _rate_from_row)
    return _dedupe_sizes(sizes), _dedupe_rates(rates)


def _parse_table(rows, required_headers, builder):
    parsed = []
    for row_index, row in enumerate(rows):
        normalized = [_normalize_header(value) for value in row]
        indexes = _header_indexes(normalized, required_headers)
        if indexes is None:
            continue
        for data_row in rows[row_index + 1:]:
            if _is_blank_row(data_row):
                continue
            try:
                parsed.append(builder(data_row, indexes))
            except ValueError:
                continue
    return parsed


def _header_indexes(normalized_row, header_aliases):
    indexes = {}
    for field, aliases in header_aliases.items():
        for alias in aliases:
            if alias in normalized_row:
                indexes[field] = normalized_row.index(alias)
                break
        if field not in indexes:
            return None
    return indexes


def _size_from_row(row, indexes) -> ParsedSize:
    tile_size = _required_text(row[indexes["tile_size"]])
    pieces_per_box = int(_required_number(row[indexes["pieces_per_box"]]))
    area_per_box = float(_required_number(row[indexes["area_per_box"]]))
    return ParsedSize(tile_size=tile_size, pieces_per_box=pieces_per_box, area_per_box=area_per_box)


def _rate_from_row(row, indexes) -> ParsedRate:
    grade = GRADE_MAP.get(_required_text(row[indexes["grade"]]))
    if not grade:
        raise ValueError("Unknown grade")
    tile_size = _required_text(row[indexes["tile_size"]])
    rate_per_meter = float(_required_number(row[indexes["rate_per_meter"]]))
    return ParsedRate(tile_size=tile_size, grade=grade, rate_per_meter=rate_per_meter)


def _dedupe_sizes(sizes: list[ParsedSize]) -> list[ParsedSize]:
    by_key = {}
    for size in sizes:
        existing = by_key.get(size.tile_size)
        if existing and existing != size:
            raise SystemExit(f"Conflicting size row for {size.tile_size}: {existing} vs {size}")
        by_key[size.tile_size] = size
    return list(by_key.values())


def _dedupe_rates(rates: list[ParsedRate]) -> list[ParsedRate]:
    by_key = {}
    for rate in rates:
        key = (rate.tile_size, rate.grade)
        existing = by_key.get(key)
        if existing and existing != rate:
            raise SystemExit(f"Conflicting rate row for {key}: {existing} vs {rate}")
        by_key[key] = rate
    return list(by_key.values())


def build_diff(db, sizes: list[ParsedSize], rates: list[ParsedRate]) -> dict:
    existing_sizes = {row.tile_size: row for row in db.scalars(select(TileSize)).all()}
    existing_rates = {(row.tile_size, row.grade): row for row in db.scalars(select(TileRate)).all()}

    size_changes = []
    for size in sizes:
        current = existing_sizes.get(size.tile_size)
        if not current:
            size_changes.append({"action": "create", **size.__dict__})
        elif current.pieces_per_box != size.pieces_per_box or float(current.area_per_box) != size.area_per_box:
            size_changes.append({
                "action": "update",
                "tile_size": size.tile_size,
                "before": {"pieces_per_box": current.pieces_per_box, "area_per_box": current.area_per_box},
                "after": {"pieces_per_box": size.pieces_per_box, "area_per_box": size.area_per_box},
            })

    rate_changes = []
    for rate in rates:
        current = existing_rates.get((rate.tile_size, rate.grade))
        if not current:
            rate_changes.append({"action": "create", **rate.__dict__})
        elif float(current.rate_per_meter) != rate.rate_per_meter:
            rate_changes.append({
                "action": "update",
                "tile_size": rate.tile_size,
                "grade": rate.grade,
                "before": {"rate_per_meter": current.rate_per_meter},
                "after": {"rate_per_meter": rate.rate_per_meter},
            })

    return {
        "sizes_found": len(sizes),
        "rates_found": len(rates),
        "size_changes": size_changes,
        "rate_changes": rate_changes,
        "unchanged_sizes": len(sizes) - len(size_changes),
        "unchanged_rates": len(rates) - len(rate_changes),
    }


def apply_changes(db, sizes: list[ParsedSize], rates: list[ParsedRate]):
    for size in sizes:
        current = db.get(TileSize, size.tile_size)
        if not current:
            db.add(TileSize(**size.__dict__, active=True))
        else:
            current.pieces_per_box = size.pieces_per_box
            current.area_per_box = size.area_per_box
            current.active = True

    db.flush()

    for rate in rates:
        current = db.scalar(
            select(TileRate).where(TileRate.tile_size == rate.tile_size, TileRate.grade == rate.grade)
        )
        if not current:
            db.add(TileRate(**rate.__dict__, active=True))
        else:
            current.rate_per_meter = rate.rate_per_meter
            current.active = True


def _normalize_header(value) -> str:
    return " ".join(str(value or "").strip().lower().replace("(m2)", "").replace("m²", "").split())


def _is_blank_row(row) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def _required_text(value) -> str:
    if value is None or str(value).strip() == "":
        raise ValueError("Missing text value")
    return str(value).strip()


def _required_number(value) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError("Missing numeric value")
    return float(value)


if __name__ == "__main__":
    main()
